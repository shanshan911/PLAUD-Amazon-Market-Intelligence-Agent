from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from .normalizers import format_number, format_percent


def write_csv(df: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(path, index=False, encoding="utf-8-sig")


def dataframe_to_markdown(df: pd.DataFrame, columns: list[str], max_rows: int = 20) -> str:
    if df.empty:
        return "_暂无数据_"
    subset = df.copy()
    for col in columns:
        if col not in subset.columns:
            subset[col] = ""
    subset = subset[columns].head(max_rows).copy()
    headers = "| " + " | ".join(columns) + " |"
    separator = "| " + " | ".join(["---"] * len(columns)) + " |"
    rows = []
    for _, row in subset.iterrows():
        rows.append("| " + " | ".join("" if pd.isna(row[col]) else str(row[col]) for col in columns) + " |")
    return "\n".join([headers, separator] + rows)


def present_brand_share(df: pd.DataFrame) -> pd.DataFrame:
    result = df.copy()
    for col in ("monthly_units_share", "monthly_revenue_share"):
        if col in result:
            result[col] = result[col].map(format_percent)
    for col in ("monthly_units", "monthly_revenue"):
        if col in result:
            result[col] = result[col].map(format_number)
    return result


def present_ai_summary(df: pd.DataFrame) -> pd.DataFrame:
    result = df.copy()
    for col in ("ai_units_share", "ai_revenue_share"):
        if col in result:
            result[col] = result[col].map(format_percent)
    for col in ("category_units", "category_revenue", "ai_competitor_units", "ai_competitor_revenue"):
        if col in result:
            result[col] = result[col].map(format_number)
    return result


def write_markdown_report(
    output_path: Path,
    week_id: str,
    brand_share: pd.DataFrame,
    ai_summary: pd.DataFrame,
    ai_detail: pd.DataFrame,
    run_log: list[dict[str, Any]],
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    brand_present = present_brand_share(brand_share)
    ai_present = present_ai_summary(ai_summary)
    ai_detail_present = ai_detail.copy()
    for col in ("monthly_units", "monthly_revenue"):
        if col in ai_detail_present:
            ai_detail_present[col] = ai_detail_present[col].map(format_number)

    content = [
        f"# PLAUD 亚马逊周度监控报告样例 - {week_id}",
        "",
        "## 运行状态",
        dataframe_to_markdown(pd.DataFrame(run_log), ["marketplace", "status", "source_file", "warnings", "error"], 30),
        "",
        "## 品牌/竞品市占",
        dataframe_to_markdown(
            brand_present,
            [
                "marketplace",
                "brand",
                "brand_group",
                "monthly_units",
                "monthly_units_share",
                "monthly_revenue",
                "monthly_revenue_share",
                "trend_status",
            ],
            50,
        ),
        "",
        "## AI 竞品汇总",
        dataframe_to_markdown(
            ai_present,
            [
                "marketplace",
                "ai_competitor_asin_count",
                "ai_competitor_units",
                "ai_units_share",
                "ai_competitor_revenue",
                "ai_revenue_share",
                "trend_status",
            ],
            20,
        ),
        "",
        "## AI 竞品 ASIN 明细 Top 20",
        dataframe_to_markdown(
            ai_detail_present.sort_values(["marketplace", "monthly_revenue"], ascending=[True, False])
            if not ai_detail_present.empty
            else ai_detail_present,
            [
                "marketplace",
                "asin",
                "standard_brand",
                "monthly_units",
                "monthly_revenue",
                "ai_matched_keywords",
                "product_title",
            ],
            20,
        ),
        "",
    ]
    output_path.write_text("\n".join(content), encoding="utf-8")


def write_excel_report(
    output_path: Path,
    week_id: str,
    brand_share: pd.DataFrame,
    ai_summary: pd.DataFrame,
    ai_detail: pd.DataFrame,
    run_log: list[dict[str, Any]],
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    summary = build_excel_summary(week_id, brand_share, ai_summary)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="汇总", index=False)
        brand_share.to_excel(writer, sheet_name="品牌市占", index=False)
        ai_summary.to_excel(writer, sheet_name="AI竞品汇总", index=False)
        ai_detail.to_excel(writer, sheet_name="AI竞品明细", index=False)
        pd.DataFrame(run_log).to_excel(writer, sheet_name="运行日志", index=False)

        for sheet in writer.book.worksheets:
            style_worksheet(sheet)


def build_excel_summary(week_id: str, brand_share: pd.DataFrame, ai_summary: pd.DataFrame) -> pd.DataFrame:
    rows: list[dict[str, Any]] = []
    marketplaces = sorted(set(brand_share.get("marketplace", pd.Series(dtype=str)).dropna().tolist()))
    for marketplace in marketplaces:
        brand_site = brand_share[brand_share["marketplace"] == marketplace]
        plaud = first_row(brand_site[brand_site["brand"] == "PLAUD"])
        competitors = first_row(brand_site[brand_site["brand"] == "COMPETITORS_TOTAL"])
        ai = first_row(ai_summary[ai_summary["marketplace"] == marketplace]) if not ai_summary.empty else {}
        rows.append(
            {
                "周次": week_id,
                "站点": marketplace,
                "PLAUD月销量": plaud.get("monthly_units"),
                "PLAUD销量占比": plaud.get("monthly_units_share"),
                "PLAUD月销售额": plaud.get("monthly_revenue"),
                "PLAUD销售额占比": plaud.get("monthly_revenue_share"),
                "竞品合计月销量": competitors.get("monthly_units"),
                "竞品合计销量占比": competitors.get("monthly_units_share"),
                "竞品合计月销售额": competitors.get("monthly_revenue"),
                "竞品合计销售额占比": competitors.get("monthly_revenue_share"),
                "AI竞品ASIN数": ai.get("ai_competitor_asin_count"),
                "AI竞品月销量": ai.get("ai_competitor_units"),
                "AI竞品销量占比": ai.get("ai_units_share"),
                "AI竞品月销售额": ai.get("ai_competitor_revenue"),
                "AI竞品销售额占比": ai.get("ai_revenue_share"),
            }
        )
    return pd.DataFrame(rows)


def first_row(df: pd.DataFrame) -> dict[str, Any]:
    if df.empty:
        return {}
    return df.iloc[0].to_dict()


def style_worksheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="EAF2F8")
    header_font = Font(bold=True, color="1F4D78")
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    sheet.freeze_panes = "A2"
    for column_cells in sheet.columns:
        values = ["" if cell.value is None else str(cell.value) for cell in column_cells]
        width = min(max(max((len(value) for value in values), default=8) + 2, 10), 42)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            header = sheet.cell(row=1, column=cell.column).value or ""
            if "占比" in str(header) or str(header).endswith("_share"):
                cell.number_format = "0.00%"
            elif "销售额" in str(header) or "revenue" in str(header):
                cell.number_format = "#,##0.00"
            elif "销量" in str(header) or "units" in str(header) or "ASIN数" in str(header):
                cell.number_format = "#,##0"


def write_snapshot(path: Path, brand_share: pd.DataFrame, ai_summary: pd.DataFrame) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    payload = {
        "brand_share": brand_share.to_dict(orient="records"),
        "ai_summary": ai_summary.to_dict(orient="records"),
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, default=str), encoding="utf-8")


def read_snapshot(path: str | Path | None) -> dict[str, pd.DataFrame] | None:
    if not path:
        return None
    snapshot_path = Path(path)
    if not snapshot_path.exists():
        return None
    payload = json.loads(snapshot_path.read_text(encoding="utf-8"))
    return {
        "brand_share": pd.DataFrame(payload.get("brand_share", [])),
        "ai_summary": pd.DataFrame(payload.get("ai_summary", [])),
    }
