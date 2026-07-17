from __future__ import annotations

import argparse
import json
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from plaud_monitor.config import DEFAULT_CONFIG, deep_merge  # noqa: E402
from plaud_monitor.normalizers import normalize_text  # noqa: E402


DEFAULT_SOURCE = "/Users/plaud/Downloads/PLAUD_监控Agent_资源准备清单_v1.2 (1).xlsx"
DEFAULT_SCOPE = ["US", "UK", "DE", "FR", "IT", "ES", "JP"]
CURRENCY_BY_MARKETPLACE = {
    "US": "USD",
    "CA": "CAD",
    "UK": "GBP",
    "DE": "EUR",
    "FR": "EUR",
    "IT": "EUR",
    "ES": "EUR",
    "JP": "JPY",
}


def clean_site(value: Any) -> str:
    text = normalize_text(value).upper()
    match = re.match(r"^\s*([A-Z]{2})\b", text)
    if match:
        return match.group(1)
    text = re.sub(r"[^A-Z]", "", text)
    return text


def split_keywords(value: Any) -> list[str]:
    text = normalize_text(value)
    if not text:
        return []
    parts = re.split(r"[,，/、]+", text)
    return [part.strip() for part in parts if part.strip()]


def first_non_empty(*values: Any) -> str:
    for value in values:
        text = normalize_text(value)
        if text:
            return text
    return ""


def read_category_rows(wb, scope: list[str] | None) -> dict[str, dict[str, str]]:
    ws = wb["3-目标类目URL"]
    result: dict[str, dict[str, str]] = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        site = clean_site(row[0])
        if not site or site in {"站点", "项目"}:
            continue
        if scope and site not in scope:
            continue
        path = normalize_text(row[1])
        url = normalize_text(row[2])
        keyword = normalize_text(row[3])
        if not any([path, url, keyword]):
            continue
        result[site] = {
            "currency": CURRENCY_BY_MARKETPLACE.get(site, ""),
            "keyword": keyword,
            "category_path": path,
            "category_url": url,
        }
    return result


def read_plaud_aliases(wb) -> list[str]:
    ws = wb["4-PLAUD品牌别名"]
    aliases: list[str] = []
    for row in ws.iter_rows(min_row=4, values_only=True):
        alias = normalize_text(row[1])
        if alias and alias not in aliases:
            aliases.append(alias)
    return aliases


def read_competitors(wb, scope: list[str] | None) -> dict[str, list[dict[str, Any]]]:
    ws = wb["5-竞品品牌清单"]
    result: dict[str, list[dict[str, Any]]] = defaultdict(list)
    seen: dict[str, set[str]] = defaultdict(set)
    current_site = ""
    for row in ws.iter_rows(min_row=4, values_only=True):
        raw_site = clean_site(row[1])
        brand = normalize_text(row[2])
        priority = normalize_text(row[3])
        note = normalize_text(row[4])
        if raw_site:
            current_site = raw_site
        site = current_site
        if not site or site in {"站点", "示例"}:
            continue
        if scope and site not in scope:
            continue
        if not brand or brand in {"竞品品牌名", "待提供"}:
            continue
        dedupe_key = brand.casefold()
        if dedupe_key in seen[site]:
            continue
        seen[site].add(dedupe_key)
        result[site].append(
            {
                "brand": brand,
                "aliases": [],
                "priority": priority or "中",
                "include_in_total": True,
                "note": note,
            }
        )
    return dict(result)


def read_ai_rules(wb, scope: list[str] | None) -> dict[str, Any]:
    ws = wb["6-AI判定规则"]
    per_site: dict[str, list[str]] = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        site = clean_site(row[0])
        if not site:
            continue
        if site == "场景":
            break
        if scope and site not in scope:
            continue
        keywords = split_keywords(row[2])
        if keywords:
            per_site[site] = keywords

    default_keywords = per_site.get("US") or per_site.get("UK") or ["AI", "A.I.", "Artificial Intelligence"]
    marketplace_keywords: dict[str, list[str]] = {}
    default_set = {item.casefold() for item in default_keywords}
    for site, keywords in per_site.items():
        extras = [keyword for keyword in keywords if keyword.casefold() not in default_set]
        if extras:
            marketplace_keywords[site] = extras
    return {
        "default_keywords": default_keywords,
        "marketplace_keywords": marketplace_keywords,
        "exclude_terms": ["MAIN"],
    }


def has_seller_sprite_account(wb) -> bool:
    ws = wb["1-卖家精灵账号"]
    values = []
    for row in ws.iter_rows(values_only=True):
        values.extend(normalize_text(cell) for cell in row if normalize_text(cell))
    joined = "\n".join(values)
    return "账号：" in joined or "账号:" in joined or bool(normalize_text(ws["B3"].value))


def read_buyer_account_sites(wb, scope: list[str] | None) -> dict[str, dict[str, Any]]:
    ws = wb["2-亚马逊买家账号"]
    result: dict[str, dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        site = clean_site(row[0])
        if not site or site in {"站点", "示例"}:
            continue
        if scope and site not in scope:
            continue
        email = normalize_text(row[1])
        postcode = normalize_text(row[5])
        if email or postcode:
            result[site] = {
                "available": bool(email),
                "postal_code": postcode,
                "note": normalize_text(row[6]),
            }
    return result


def build_missing_report(
    scope: list[str],
    categories: dict[str, dict[str, str]],
    competitors: dict[str, list[dict[str, Any]]],
    plaud_aliases: list[str],
    buyer_accounts: dict[str, dict[str, Any]],
    seller_account_available: bool,
) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    rows.append(
        {
            "priority": "P0",
            "item": "卖家精灵账号 + 导出权限",
            "status": "已提供" if seller_account_available else "缺失",
            "note": "已检测到账号信息；敏感凭据未写入开发配置。"
            if seller_account_available
            else "Sheet 1 未检测到账号信息。",
        }
    )
    rows.append(
        {
            "priority": "P0",
            "item": "PLAUD 品牌别名清单",
            "status": "已提供" if plaud_aliases else "缺失",
            "note": f"{len(plaud_aliases)} 个别名。",
        }
    )
    for site in scope:
        category = categories.get(site, {})
        rows.append(
            {
                "priority": "P0",
                "item": f"{site} 目标 BSR 类目 URL",
                "status": "已提供" if category.get("category_url") else "缺失",
                "note": category.get("category_url", ""),
            }
        )
        rows.append(
            {
                "priority": "P0",
                "item": f"{site} 亚马逊买家账号",
                "status": "已提供" if buyer_accounts.get(site, {}).get("available") else "缺失",
                "note": "仅写入可用状态和邮编，不写入账号邮箱。",
            }
        )
        rows.append(
            {
                "priority": "P0",
                "item": f"{site} 核心竞品品牌",
                "status": "已提供" if competitors.get(site) else "缺失",
                "note": f"{len(competitors.get(site, []))} 个品牌。",
            }
        )
    return rows


def write_markdown_report(path: Path, rows: list[dict[str, str]], config: dict[str, Any]) -> None:
    lines = [
        "# P0 资源导入报告",
        "",
        "说明：账号密码、2FA、Cookie、买家账号邮箱等敏感信息不会写入开发配置。",
        "",
        "## 配置摘要",
        "",
        f"- 监控站点：{', '.join(config['monitoring']['marketplaces'])}",
        f"- 类目配置站点数：{len(config.get('marketplaces', {}))}",
        f"- PLAUD 别名数：{len(config.get('plaud', {}).get('aliases', []))}",
        f"- AI 关键词站点数：{len(config.get('ai_rules', {}).get('marketplace_keywords', {})) + 1}",
        "",
        "## P0 状态",
        "",
        "| 优先级 | 资源项 | 状态 | 备注 |",
        "| --- | --- | --- | --- |",
    ]
    for row in rows:
        lines.append(f"| {row['priority']} | {row['item']} | {row['status']} | {row['note']} |")
    lines.append("")
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="Import P0 resources from PLAUD resource checklist")
    parser.add_argument("--file", default=DEFAULT_SOURCE, help="Resource checklist Excel file")
    parser.add_argument("--output-config", default="config/monitor_config.p0.json", help="Generated config path")
    parser.add_argument("--report", default="outputs/resource_import/p0_resource_status.md", help="Generated report path")
    parser.add_argument(
        "--scope",
        default=",".join(DEFAULT_SCOPE),
        help="Comma-separated marketplaces to import. Use ALL to import every marketplace in the file.",
    )
    args = parser.parse_args()

    scope = None if args.scope.upper() == "ALL" else [clean_site(item) for item in args.scope.split(",") if clean_site(item)]
    wb = load_workbook(args.file, data_only=True)

    categories = read_category_rows(wb, scope)
    competitors = read_competitors(wb, scope)
    plaud_aliases = read_plaud_aliases(wb)
    ai_rules = read_ai_rules(wb, scope)
    buyer_accounts = read_buyer_account_sites(wb, scope)
    seller_account_available = has_seller_sprite_account(wb)

    marketplaces = scope or sorted(categories)
    config_override: dict[str, Any] = {
        "monitoring": {
            "week_id": "2026-W20",
            "marketplaces": marketplaces,
        },
        "marketplaces": categories,
        "plaud": {"aliases": plaud_aliases},
        "competitors": competitors,
        "ai_rules": ai_rules,
        "input": {
            "raw_dir": "data/raw",
            "file_pattern": "{week_id}_{marketplace}_*.xlsx",
        },
        "resource_import": {
            "source_file": str(Path(args.file).resolve()),
            "seller_sprite_account_available": seller_account_available,
            "amazon_access": buyer_accounts,
            "sensitive_fields_excluded": True,
        },
    }
    config = deep_merge(DEFAULT_CONFIG, config_override)

    output_config = Path(args.output_config)
    output_config.parent.mkdir(parents=True, exist_ok=True)
    output_config.write_text(json.dumps(config, ensure_ascii=False, indent=2), encoding="utf-8")

    rows = build_missing_report(
        marketplaces,
        categories,
        competitors,
        plaud_aliases,
        buyer_accounts,
        seller_account_available,
    )
    write_markdown_report(Path(args.report), rows, config)

    print(f"Generated config: {output_config}")
    print(f"Generated report: {args.report}")
    print(f"Marketplaces: {', '.join(marketplaces)}")
    print("Competitors:", ", ".join(f"{site}={len(items)}" for site, items in competitors.items()))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
